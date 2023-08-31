from __future__ import absolute_import, division, print_function
import argparse
from copy import deepcopy
import os
import os.path as op
import code
import json
import time
import datetime
import torch
import torchvision.models as models
import gc
import numpy as np
import cv2
from openpyxl import load_workbook, Workbook
from torchvision.utils import make_grid
from src.datasets.utils import copy2cpu, crop_image, project_pcl, trans_mat_2_dict, INTRINSIC
from torch.utils.tensorboard import SummaryWriter

from src.modeling.bert import BertConfig, Graphormer
import src.modeling.model as Models
from src.modeling._smpl import SMPL, SMPLX, SMPLH36M, Mesh, SMPLXMesh
from src.modeling.hrnet.hrnet_cls_net_gridfeat import get_cls_net_gridfeat
from src.modeling.hrnet.config import config as hrnet_config
from src.modeling.hrnet.config import update_config as hrnet_update_config
from src.datasets.build import make_data_loader
from src.modeling.pointnet2.pointnet2_modules import PointnetSAModule
from src.utils.geometric_layers import orthographic_projection
import src.modeling.data.config as cfg

from src.utils.logger import setup_logger
from src.utils.comm import synchronize, is_main_process, get_rank, all_gather
from src.utils.miscellaneous import mkdir, set_seed
from src.utils.metric_logger import AverageMeter, EvalMetricsLogger
from src.utils.metric_pampjpe import reconstruction_error
import src.datasets.fusion_dataset as Datasets
from src.utils.loss import LossManager
from src.utils.renderer import Renderer, visualize_reconstruction, visualize_reconstruction_no_text
from src.utils.visualization import EvaluateStreamPlot, O3DStreamPlot, pcl2box, pcl2sphere
from azureml.core.run import Run
from src.utils.dingtalk import TimerBot

aml_run = Run.get_context()

def save_checkpoint(model, args, epoch, iteration, num_trial=10):
    checkpoint_dir = op.join(args.output_dir, 'checkpoint',
        'checkpoint-{}-{}'.format(epoch, iteration))
    if not is_main_process():
        return checkpoint_dir
    mkdir(checkpoint_dir)
    model_to_save = model.module if hasattr(model, 'module') else model
    for i in range(num_trial):
        try:
            torch.save(model_to_save.state_dict(), op.join(checkpoint_dir, 'state_dict.bin'))
            torch.save(args, op.join(checkpoint_dir, 'training_args.bin'))
            torch.save(model_to_save, op.join(checkpoint_dir, 'model.bin'))
            logger.info("Save checkpoint to {}".format(checkpoint_dir))
            break
        except:
            pass
    else:
        logger.info("Failed to save checkpoint after {} trails.".format(num_trial))

def adjust_learning_rate(optimizer, epoch, args):
    """
    Sets the learning rate to the initial LR decayed by x every y epochs
    x = 0.1, y = args.num_train_epochs/2.0 = 100
    """
    lr = args.lr * (0.1 ** (epoch // (args.num_train_epochs/2.0)  ))
    for param_group in optimizer.param_groups:
        param_group['lr'] = lr

def mean_per_joint_position_error(pred, gt):
    """ 
    Compute mPJPE
    """
    with torch.no_grad():
        gt_pelvis = (gt[:, 2,:] + gt[:, 3,:]) / 2
        gt = gt - gt_pelvis[:, None, :]
        pred_pelvis = (pred[:, 2,:] + pred[:, 3,:]) / 2
        pred = pred - pred_pelvis[:, None, :]
        error = torch.sqrt( ((pred - gt) ** 2).sum(dim=-1)).mean(dim=-1).cpu().numpy()
        return error

def mean_per_vertex_error(pred, gt):
    """
    Compute mPVE
    """
    with torch.no_grad():
        error = torch.sqrt( ((pred - gt) ** 2).sum(dim=-1)).mean(dim=-1).cpu().numpy()
        return error

def keypoint_2d_loss(criterion_keypoints, pred_keypoints_2d, gt_keypoints_2d):
    """
    Compute 2D reprojection loss if 2D keypoint annotations are available.
    The confidence (conf) is binary and indicates whether the keypoints exist or not.
    """
    conf = 1.
    if gt_keypoints_2d.shape[2] == 3:
        conf = gt_keypoints_2d[:, :, -1].unsqueeze(-1).clone()
        gt_keypoints_2d = gt_keypoints_2d[:, :, :-1]
    loss = (conf * criterion_keypoints(pred_keypoints_2d, gt_keypoints_2d)).mean()
    return loss

def keypoint_3d_loss(criterion_keypoints, pred_keypoints_3d, gt_keypoints_3d, device):
    """
    Compute 3D keypoint loss if 3D keypoint annotations are available.
    """
    conf = 1.
    if gt_keypoints_3d.shape[2] == 4:
        conf = gt_keypoints_3d[:, :, -1].unsqueeze(-1).clone()
        gt_keypoints_3d = gt_keypoints_3d[:, :, :-1]
    if len(gt_keypoints_3d) > 0:
        return (conf * criterion_keypoints(pred_keypoints_3d, gt_keypoints_3d)).mean()
    else:
        return torch.FloatTensor(1).fill_(0.).to(device) 

def vertices_loss(criterion_vertices, pred_vertices, gt_vertices, device):
    """
    Compute per-vertex loss if vertex annotations are available.
    """
    if len(gt_vertices) > 0:
        return criterion_vertices(pred_vertices, gt_vertices)
    else:
        return torch.FloatTensor(1).fill_(0.).to(device) 

def show_att_weight(args, data_dict, cluster_centers, attention, pred_joints, pred_verts_sub2, gt_pelvis, idx):
    trans_mat = {}
    for k, v in data_dict['trans_mat'].items():
        trans_mat[k] = trans_mat_2_dict(v[idx])
    heads_num, vertex_num, _ = attention.shape
    all_head = np.zeros((vertex_num, vertex_num))
    # compute avg attention for all attention heads
    for h in range(heads_num):
        att_per_img = attention[h]
        all_head = all_head + att_per_img
    all_head = all_head / heads_num
    col_sums = all_head.sum(axis=0)
    all_head = all_head / col_sums[np.newaxis, :]

    # select the joint to show attention
    ref_joint = pred_joints[args.joint_id]
    attention_to_show = all_head[args.joint_id][22:]
    # nomalize attention
    min_v = np.min(attention_to_show)
    max_v = np.max(attention_to_show)
    norm_attention_to_show = (attention_to_show - min_v)/(max_v-min_v)

    # get verts attention lines
    verts_to_show = [ref_joint,]
    verts_lines = []
    verts_line_colors = []
    thres = 0.1
    for i in range(pred_verts_sub2.shape[0]):
        if norm_attention_to_show[i] > thres:
            verts_to_show.append(pred_verts_sub2[i])
            verts_lines.append([0, len(verts_to_show)-1])
            verts_line_colors.append(1 + thres - norm_attention_to_show[i] * np.asarray((200, 208, 255)) / 255)

    def draw_pcl_att_map(pcl_input):
        if pcl_input is 'radar0':
            image = 'image0'
        else:
            image = 'image{}'.format(pcl_input[-1])
        # project points to the image
        orig_img = data_dict['orig_img'][image][idx]
        pcl_2d = project_pcl(data_dict[pcl_input][idx][:,:3] + gt_pelvis, trans_mat[image], intrinsic=INTRINSIC[image])
        # project cluster centers to the image
        center_2d = project_pcl(cluster_centers[pcl_input][idx] + gt_pelvis, trans_mat[image], intrinsic=INTRINSIC[image])
        blank = np.zeros_like(orig_img[:,:,:1])
        start_idx = 655 + 49 * args.inputs.index(pcl_input)
        # sort the weight
        weight_sorted_idx = np.argsort(norm_attention_to_show[[i+start_idx for i in range(center_2d.shape[0])]])
        # plot cluster centers in the image
        for i in weight_sorted_idx:
            cv2.circle(blank, (center_2d[i, 1], center_2d[i, 0]), 60, norm_attention_to_show[i+start_idx]*255, -1)
        # apply jet color map
        blank = cv2.applyColorMap(blank, 2)
        img_with_cluster = cv2.addWeighted(orig_img, 0.2, blank, 0.5, 1)
        for i in range(-1, 2):
            for j in range(-1, 2):
                img_with_cluster[pcl_2d[:,0]+i, pcl_2d[:,1]+j] = [0, 255, 0]
        # crop the image area of people
        img_with_cluster = crop_image(pred_joints+gt_pelvis, img_with_cluster, trans_mat[image], margin=0.3, square=True, intrinsic=INTRINSIC[image])
        img_with_cluster = cv2.resize(img_with_cluster, (224, 224))
        return img_with_cluster
    
    def draw_img_att_map(img_input):
        # crop the image area of people
        croped_img = crop_image(joints=pred_joints+gt_pelvis, image=data_dict['orig_img'][img_input][idx], 
                                trans_mat=trans_mat[img_input], margin=0.3, square=True, intrinsic=INTRINSIC[img_input])
        croped_img = cv2.resize(croped_img, (224, 224))
        grid_width = 224 // 7
        blank = np.zeros((224, 224, 1), dtype=np.uint8)
        start_idx = 655 + 49 * args.inputs.index(img_input)
        # plot grid
        for i in range(7):
            for j in range(7):
                cv2.rectangle(blank, (i*grid_width,j*grid_width), ((i+1)*grid_width,(j+1)*grid_width), 
                            norm_attention_to_show[i*7+j+start_idx]*255, -1)
        # apply jet color map
        blank = cv2.applyColorMap(blank, 2)
        img_with_grid = cv2.addWeighted(croped_img, 0.5, blank, 0.5, 1)
        return img_with_grid
    
    img_to_show = []
    # draw image attention map
    for img_input in args.input_dict['image']:
        if img_input in args.inputs:
            img_to_show.append(draw_img_att_map(img_input))
    
    # draw point attention map
    for pcl_input in args.input_dict['depth'] + args.input_dict['radar']:
        if pcl_input in args.inputs:
            img_to_show.append(draw_pcl_att_map(pcl_input))
            
    heatmap = np.hstack(img_to_show)
    cv2.imwrite('heatmap.png', heatmap)
    cv2.namedWindow('heatmap', 0)
    cv2.resizeWindow("heatmap", 1280, 480)
    cv2.imshow('heatmap', heatmap)
    cv2.waitKey(1)

    res = dict(
        vert_pcl=dict(pcl=pred_verts_sub2, color=np.asarray([255, 181, 74])/255,),
        vert_lines=dict(pcl=verts_to_show, lines=np.asarray(verts_lines), colors=verts_line_colors),
    )

    return res

def visualize_mesh(renderer, images, gt_keypoints_2d, pred_vertices, pred_camera, pred_keypoints_2d):
    """Tensorboard logging."""
    gt_keypoints_2d = gt_keypoints_2d.cpu().numpy()
    to_lsp = list(range(14))
    rend_imgs = []
    batch_size = pred_vertices.shape[0]
    # Do visualization for the first 6 images of the batch
    for i in range(min(batch_size, 10)):
        img = images[i].cpu().numpy().transpose(1,2,0)
        # Get LSP keypoints from the full list of keypoints
        gt_keypoints_2d_ = gt_keypoints_2d[i, to_lsp]
        pred_keypoints_2d_ = pred_keypoints_2d.cpu().numpy()[i, to_lsp]
        # Get predict vertices for the particular example
        vertices = pred_vertices[i].cpu().numpy()
        cam = pred_camera[i].cpu().numpy()
        # Visualize reconstruction and detected pose
        rend_img = visualize_reconstruction(img, 224, gt_keypoints_2d_, vertices, pred_keypoints_2d_, cam, renderer)
        rend_img = rend_img.transpose(2,0,1)
        rend_imgs.append(torch.from_numpy(rend_img))   
    rend_imgs = make_grid(rend_imgs, nrow=1)
    return rend_imgs

def visualize_mesh_test(renderer, images, pred_vertices, pred_camera):
    img = images.cpu().numpy().transpose(1,2,0)
    # Get predict vertices for the particular example
    vertices_full = pred_vertices.cpu().numpy() 
    cam = pred_camera.cpu().numpy()
    # Visualize only mesh reconstruction 
    rend_img = visualize_reconstruction_no_text(img, 224, vertices_full, cam, renderer, color='light_blue')
    return rend_img

def run_train(args, train_dataloader, val_dataloader, fusiuon_model, smpl, mesh_sampler, start_epoch):
    smpl.eval()
    max_iter = len(train_dataloader)
    iters_per_epoch = max_iter // args.num_train_epochs
    start_iter = start_epoch * iters_per_epoch
    if iters_per_epoch < 1000:
        args.logging_steps = 500

    optimizer = torch.optim.Adam(params=list(fusiuon_model.parameters()),
                                           lr=args.lr,
                                           betas=(0.9, 0.999),
                                           weight_decay=0)

    # define loss function (criterion) and optimizer
    criterion_3d_keypoints = torch.nn.MSELoss(reduction='none').to(args.device)
    criterion_vertices = torch.nn.L1Loss().to(args.device)
    criterion_2d_keypoints = torch.nn.MSELoss(reduction='none').cuda(args.device)

    if args.distributed:
        fusiuon_model = torch.nn.parallel.DistributedDataParallel(
            fusiuon_model, device_ids=[args.local_rank], 
            output_device=args.local_rank,
            find_unused_parameters=True,
        )

        logger.info(
                ' '.join(
                ['Local rank: {o}', 'Max iteration: {a}', 'iters_per_epoch: {b}','num_train_epochs: {c}',]
                ).format(o=args.local_rank, a=max_iter, b=iters_per_epoch, c=args.num_train_epochs)
            )

    TIMESTAMP = "{0:%Y-%m-%dT%H-%M-%S/}".format(datetime.datetime.now())
    loss_manager_train = SummaryWriter(args.output_dir + '/loss/train/' + TIMESTAMP)
    loss_manager_eval = SummaryWriter(args.output_dir + '/loss/eval/' + TIMESTAMP)
    
    renderer = Renderer(faces=smpl.faces.cpu().numpy())
    
    start_training_time = time.time()
    end = time.time()
    fusiuon_model.train()
    batch_time = AverageMeter()
    data_time = AverageMeter()
    log_losses = AverageMeter()
    log_loss_3djoints = AverageMeter()
    log_loss_2djoints = AverageMeter()
    log_loss_vertices = AverageMeter()
    log_eval_metrics = EvalMetricsLogger()

    for iteration, data_dict in enumerate(train_dataloader):
        # gc.collect()
        # torch.cuda.empty_cache()
        fusiuon_model.train()
        iteration += start_iter + 1
        epoch = iteration // iters_per_epoch
        adjust_learning_rate(optimizer, epoch, args)
        data_time.update(time.time() - end)

        # normalize gt joints
        gt_3d_joints = data_dict['joints_3d'].to(args.device)
        gt_3d_pelvis = data_dict['root_pelvis'].to(args.device)[:,None,:3]
        gt_3d_joints[:,:,:3] -= gt_3d_pelvis
        batch_size = gt_3d_joints.shape[0]

        joint_mask = data_dict['joint_mask'].to(args.device)
        vert_mask = data_dict['vert_mask'].to(args.device)

        # generate simplified mesh
        gt_vertices = data_dict['vertices'].to(args.device)
        gt_vertices_sub2 = mesh_sampler.downsample(gt_vertices, n1=0, n2=2)
        gt_vertices_sub = mesh_sampler.downsample(gt_vertices)

        # normalize gt based on smpl's pelvis
        gt_vertices_sub2 -= gt_3d_pelvis
        gt_vertices_sub -= gt_3d_pelvis
        gt_vertices -= gt_3d_pelvis
        
        for m in args.inputs:
            data_dict[m] = data_dict[m].to(args.device)

        # prepare masks for mask vertex/joint modeling
        joint_mask_ = joint_mask.expand(-1,-1,2051)
        vert_mask_ = vert_mask.expand(-1,-1,2051)
        meta_masks = torch.cat([joint_mask_, vert_mask_], dim=1)

        # forward-pass
        pred_dict, pred_3d_joints, pred_vertices_sub2, pred_vertices_sub, pred_vertices = fusiuon_model(args, data_dict, smpl, mesh_sampler, meta_masks=meta_masks, is_train=True)

        # compute 3d joint loss  (where the joints are directly output from transformer)
        loss_3d_joints = keypoint_3d_loss(criterion_3d_keypoints, pred_3d_joints, gt_3d_joints, args.device)
        # compute 3d vertex loss
        loss_vertices = (args.vloss_w_sub2 * vertices_loss(criterion_vertices, pred_vertices_sub2, gt_vertices_sub2, args.device) + \
                            args.vloss_w_sub * vertices_loss(criterion_vertices, pred_vertices_sub, gt_vertices_sub, args.device) + \
                            args.vloss_w_full * vertices_loss(criterion_vertices, pred_vertices, gt_vertices, args.device))

        # we empirically use hyperparameters to balance difference losses
        loss = args.joints_loss_weight*loss_3d_joints + args.vertices_loss_weight*loss_vertices
        
        if args.joints_2d_loss:
            gt_2d_joints = data_dict['joints_2d'][args.joints_2d_loss].float().to(args.device)
            pred_2d_joints = orthographic_projection(pred_3d_joints, pred_dict['camera'])
            loss_2d_joints = keypoint_2d_loss(criterion_2d_keypoints, pred_2d_joints, gt_2d_joints)
            log_loss_2djoints.update(loss_2d_joints.item(), batch_size)
            loss_manager_train.add_scalar("2d_joint_loss", loss_2d_joints.item(), iteration)
            loss += args.vertices_loss_weight*loss_2d_joints
        
        if args.model == 'TokenFusion':
            score_loss = torch.sum(torch.abs(pred_dict.get('pred_score', torch.zeros(1).cuda())))
            loss_manager_train.add_scalar("score_loss", score_loss.item(), iteration)
            loss += 1e-3*score_loss
        
        loss_manager_train.add_scalar("3d_joint_loss", loss_3d_joints.item(), iteration)
        loss_manager_train.add_scalar("vertex_loss", loss_vertices.item(), iteration)
        loss_manager_train.add_scalar("total_loss", loss.item(), iteration)
        
        # update logs
        log_loss_3djoints.update(loss_3d_joints.item(), batch_size)
        log_loss_vertices.update(loss_vertices.item(), batch_size)
        log_losses.update(loss.item(), batch_size)

        # back prop
        optimizer.zero_grad()
        loss.backward() 
        optimizer.step()

        batch_time.update(time.time() - end)
        end = time.time()

        if iteration % args.logging_steps == 0 or iteration == max_iter:
            eta_seconds = batch_time.avg * (max_iter - iteration)
            eta_string = str(datetime.timedelta(seconds=int(eta_seconds)))
            logger.info(
                ' '.join(
                ['eta: {eta}', 'epoch: {ep}', 'iter: {iter}', 'max mem : {memory:.0f}',]
                ).format(eta=eta_string, ep=epoch, iter=iteration, 
                    memory=torch.cuda.max_memory_allocated() / 1024.0 / 1024.0) 
                + '  loss: {:.4f}, 3d joint loss: {:.4f}, vertex loss: {:.4f}, compute: {:.4f}, data: {:.4f}, lr: {:.6f}'.format(
                    log_losses.avg, log_loss_3djoints.avg, log_loss_vertices.avg, batch_time.avg, data_time.avg, 
                    optimizer.param_groups[0]['lr'])
            )

            aml_run.log(name='Loss', value=float(log_losses.avg))
            aml_run.log(name='3d joint Loss', value=float(log_loss_3djoints.avg))
            aml_run.log(name='vertex Loss', value=float(log_loss_vertices.avg))

            if args.joints_2d_loss:
                aml_run.log(name='2d joint Loss', value=float(log_loss_2djoints.avg))
                visual_imgs = visualize_mesh(renderer,
                                            data_dict['orig_img'][args.joints_2d_loss].detach(),
                                            data_dict['joints_2d'][args.joints_2d_loss][:,:,:2].detach(),
                                            pred_vertices.detach(), 
                                            pred_dict['camera'].detach(),
                                            pred_2d_joints.detach())
                visual_imgs = visual_imgs.permute(1,2,0).numpy()
                
                if is_main_process():
                    if not os.path.exists(os.path.join(args.output_dir, 'visual')):
                        os.makedirs(os.path.join(args.output_dir, 'visual'))
                    stamp = str(epoch) + '_' + str(iteration)
                    temp_fname = os.path.join(args.output_dir, 'visual', 'visual_{}.jpg'.format(stamp))
                    cv2.imwrite(temp_fname, np.asarray(visual_imgs[:,:,::-1]*255))
                    aml_run.log_image(name='visual results', path=temp_fname)
                
        if iteration % iters_per_epoch == 0:
            val_mPVE, val_mPJPE, val_PAmPJPE, val_count = run_validate(args, val_dataloader, 
                                                fusiuon_model, 
                                                epoch, 
                                                smpl,
                                                mesh_sampler,
                                                )
            
            aml_run.log(name='mPVE', value=float(1000*val_mPVE))
            aml_run.log(name='mPJPE', value=float(1000*val_mPJPE))
            aml_run.log(name='PAmPJPE', value=float(1000*val_PAmPJPE))
            loss_manager_eval.add_scalar("mPVE", float(1000*val_mPVE), epoch)
            loss_manager_eval.add_scalar("mPJPE", float(1000*val_mPJPE), epoch)
            loss_manager_eval.add_scalar("PAmPJPE", float(1000*val_PAmPJPE), epoch)
            logger.info(
                ' '.join(['Validation', 'epoch: {ep}',]).format(ep=epoch) 
                + '  mPVE: {:6.2f}, mPJPE: {:6.2f}, PAmPJPE: {:6.2f}, Data Count: {:6.2f}'.format(1000*val_mPVE, 1000*val_mPJPE, 1000*val_PAmPJPE, val_count)
            )

            if val_PAmPJPE<log_eval_metrics.PAmPJPE or epoch%10==0:
                save_checkpoint(fusiuon_model, args, epoch, iteration)
                log_eval_metrics.update(val_mPVE, val_mPJPE, val_PAmPJPE, epoch)
        
    total_training_time = time.time() - start_training_time
    total_time_str = str(datetime.timedelta(seconds=total_training_time))
    logger.info('Total training time: {} ({:.4f} s / iter)'.format(
        total_time_str, total_training_time / max_iter)
    )
    save_checkpoint(fusiuon_model, args, epoch, iteration)

    logger.info(
        ' Best Results:'
        + '  mPVE: {:6.2f}, mPJPE: {:6.2f}, PAmPJPE: {:6.2f}, at epoch {:6.2f}'.format(1000*log_eval_metrics.mPVE, 1000*log_eval_metrics.mPJPE, 1000*log_eval_metrics.PAmPJPE, log_eval_metrics.epoch)
    )


def run_eval_general(args, val_dataloader, fusion_model, smpl, mesh_sampler):
    smpl.eval()
    epoch = 0
    if args.distributed:
        fusion_model = torch.nn.parallel.DistributedDataParallel(
            fusion_model, device_ids=[args.local_rank], 
            output_device=args.local_rank,
            find_unused_parameters=True,
        )
    fusion_model.eval()

    val_mPVE, val_mPJPE, val_PAmPJPE, val_count = run_validate(args, val_dataloader, fusion_model, epoch, smpl, mesh_sampler)

    aml_run.log(name='mPVE', value=float(1000*val_mPVE))
    aml_run.log(name='mPJPE', value=float(1000*val_mPJPE))
    aml_run.log(name='PAmPJPE', value=float(1000*val_PAmPJPE))

    logger.info(
        ' '.join(['Validation', 'epoch: {ep}',]).format(ep=epoch) 
        + '  mPVE: {:6.2f}, mPJPE: {:6.2f}, PAmPJPE: {:6.2f} '.format(1000*val_mPVE, 1000*val_mPJPE, 1000*val_PAmPJPE)
    )
    return

def result_generator(args, data_dict, pred_vertices, gt_vertices, face, attention=None, 
                     cluster_centers=None, pred_verts_sub2=None, pred_joints=None, gt_pelvis=None):
    for k, v in data_dict.items():
        if isinstance(v, torch.Tensor):
            data_dict[k] = copy2cpu(v)
    for i in range(len(pred_vertices)):
        print('frame_id:', i)
        if not args.show_att:
            depth_colors = dict(
                depth0=np.asarray([245, 157, 86]) / 255,
                depth1=np.asarray([59, 170, 235]) / 255,
                depth2=np.asarray([0, 204, 0]) / 255,
                depth3=np.asarray([202, 139, 245]) / 255,
            )
            res = {}
            res.update({
                depth:dict(
                    pcl = data_dict[depth][i][:,:3],
                    color = depth_colors[depth],
                ) for depth in args.input_dict['depth']
            })
            res.update(dict(
                radar0 = dict(
                    # mesh = pcl2box(data_dict['radar0'][i][:,:3]) if 'radar0' in args.inputs else None,
                    pcl = data_dict['radar0'][i][:,:3] if 'radar0' in args.inputs else None,
                    color = [0,0.8,0],
                ),
                pred_smpl = dict(
                    mesh = [copy2cpu(pred_vertices)[i], copy2cpu(face)],
                    color = np.asarray([179, 230, 213]) / 255
                ),
                label_smpl = dict(
                    mesh = [copy2cpu(gt_vertices)[i], copy2cpu(face)],
                    color = np.asarray([235, 189, 191]) / 255,
                ),
            ))
        else:
            for k, v in cluster_centers.items():
                cluster_centers[k] = copy2cpu(v)
            for k, v in data_dict['orig_img'].items():
                data_dict['orig_img'].update({k: copy2cpu(v)})
            res = show_att_weight(args, data_dict, cluster_centers, copy2cpu(attention)[i], copy2cpu(pred_joints)[i], copy2cpu(pred_verts_sub2)[i], copy2cpu(gt_pelvis)[i], i)
        O3DStreamPlot.pause = args.pause_at_start
        yield res

def run_validate(args, val_loader, fusion_model, epoch, smpl, mesh_sampler):
    batch_time = AverageMeter()
    mPVE = AverageMeter()
    mPJPE = AverageMeter()
    PAmPJPE = AverageMeter()

    per_joint_err = []
    per_vertex_err = []
    per_pampjpe = []

    if not args.train:
        snapshot_path = os.path.join(args.output_dir, 'snapshot', args.test_scene)
        if not os.path.exists(snapshot_path):
            os.makedirs(snapshot_path)
        plot = EvaluateStreamPlot(save_path=snapshot_path)
        renderer = Renderer(faces=smpl.faces.cpu().numpy())
    # switch to evaluate mode
    fusion_model.eval()
    smpl.eval()
    batch = 0
    pelvis = 14 if args.dataset == 'Human36MDataset' else 0
    with torch.no_grad():
        for i, data_dict in enumerate(val_loader):
            # compute output
            gt_3d_joints = data_dict['joints_3d'].to(args.device)
            gt_3d_pelvis = data_dict['root_pelvis'].to(args.device)[:,None,:3]
            gt_3d_joints[:,:,:3] -= gt_3d_pelvis

            # generate simplified mesh
            gt_vertices = data_dict['vertices'].to(args.device)
            gt_vertices_sub = mesh_sampler.downsample(gt_vertices)
            gt_vertices_sub2 = mesh_sampler.downsample(gt_vertices_sub, n1=1, n2=2)

            # normalize gt based on smpl pelvis 
            gt_vertices_sub2 -= gt_3d_pelvis
            gt_vertices_sub -= gt_3d_pelvis
            gt_vertices -= gt_3d_pelvis
                    
            for m in args.inputs:
                data_dict[m] = data_dict[m].to(args.device)
                
            # forward-pass
            pred_dict, pred_3d_joints, pred_verts_sub2, pred_vertices_sub, pred_vertices = fusion_model(args, data_dict, smpl, mesh_sampler)

            pred_3d_pelvis = pred_3d_joints[:,0,:]
            if args.dataset == 'Human36MDataset':
                gt_3d_joints = gt_3d_joints[:,cfg.J24_TO_J14,:]
                pred_3d_pelvis = smpl.get_h36m_joints(pred_vertices)[:,cfg.H36M_J17_NAME.index('Pelvis'),:]

            pred_vertices = pred_vertices - pred_3d_pelvis[:, None, :]
            pred_3d_joints = pred_3d_joints - pred_3d_pelvis[:, None, :]
            # measure errors
            error_vertices = mean_per_vertex_error(pred_vertices, gt_vertices)
            error_joints = mean_per_joint_position_error(pred_3d_joints, gt_3d_joints[:,:,:3])
            error_joints_pa = reconstruction_error(pred_3d_joints.cpu().numpy(), gt_3d_joints[:,:,:3].cpu().numpy(), reduction=None)

            per_joint_err.append(torch.norm((pred_3d_joints - gt_3d_joints[:,:,:3]), dim=-1))
            per_vertex_err.append(torch.norm((pred_vertices - gt_vertices[:,:,:3]), dim=-1))
            per_pampjpe.append(error_joints_pa)

            if not args.train and args.joints_2d_loss and i%100==0:
                # visual_imgs = visualize_mesh_test(renderer,
                #                             data_dict['orig_img'][args.joints_2d_loss][0].detach(),
                #                             pred_vertices[0].detach(), 
                #                             pred_dict['camera'][0].detach(),)
                pred_2d_joints = orthographic_projection(pred_3d_joints, pred_dict['camera'])
                visual_imgs_output = visualize_mesh(renderer,
                                                    data_dict['orig_img'][args.joints_2d_loss].detach(),
                                                    data_dict['joints_2d'][args.joints_2d_loss][:,:,:2].detach(),
                                                    pred_vertices.detach(), 
                                                    pred_dict['camera'].detach(),
                                                    pred_2d_joints.detach())
                visual_imgs = visual_imgs_output.permute(1,2,0).numpy()
                temp_fname = os.path.join(args.output_dir, 'snapshot', 'pred_{}.jpg'.format(i))
                print('save to ', temp_fname)
                cv2.imwrite(temp_fname, np.asarray(visual_imgs[:,:,::-1]*255))
                
            if len(error_vertices)>0:
                mPVE.update(np.mean(error_vertices))
            if len(error_joints)>0:
                mPJPE.update(np.mean(error_joints))
            if len(error_joints_pa)>0:
                PAmPJPE.update(np.mean(error_joints_pa))

            if args.visual and not args.train:
                attention = pred_dict.get('attention', None)
                cluster_centers = pred_dict.get('cluster_centers', None)
                gen = result_generator(args, data_dict, pred_vertices, gt_vertices, smpl.f, attention, 
                                       cluster_centers, pred_verts_sub2, pred_3d_joints, gt_3d_pelvis)
                batch += 1
                if batch > args.max_num_batch:
                    return 0, 0, 0, 0
                if args.save_snapshot:
                    print('batch_id:', i)
                    plot.show(gen, save_path=snapshot_path)
                else:
                    print('batch_id:', i)
                    plot.show(gen)

    val_mPVE = all_gather(float(mPVE.avg))
    val_mPVE = sum(val_mPVE)/len(val_mPVE)
    val_mPJPE = all_gather(float(mPJPE.avg))
    val_mPJPE = sum(val_mPJPE)/len(val_mPJPE)

    val_PAmPJPE = all_gather(float(PAmPJPE.avg))
    val_PAmPJPE = sum(val_PAmPJPE)/len(val_PAmPJPE)

    val_count = all_gather(float(mPVE.count))
    val_count = sum(val_count)

    # save error
    if not args.train:
        output_path = os.path.join(args.output_dir, 'error', args.test_scene)
        if not os.path.isdir(output_path):
            os.makedirs(output_path)
        j_err = copy2cpu(torch.vstack(per_joint_err))
        v_err = copy2cpu(torch.vstack(per_vertex_err))
        pa_err = np.hstack(per_pampjpe)
        np.save(os.path.join(output_path, "per_joint_err"), j_err)
        np.save(os.path.join(output_path, "per_vertex_err"), v_err)
        print("mean joint err (cm):", np.mean(j_err)*100)
        print("mean vertex err (cm):", np.mean(v_err)*100)
        print("max joint err (cm):", np.mean(np.max(j_err, axis=1), axis=0)*100)
        print("max vertex err (cm):", np.mean(np.max(v_err, axis=1), axis=0)*100)
        with open(os.path.join(output_path, "error.txt"), 'w') as f:
            f.write("mean joint error: " + str(np.mean(j_err)*100))
            f.write("\nmean vertex error: " + str(np.mean(v_err)*100))
            f.write("\nmax joint error: " + str(np.mean(np.max(j_err, axis=1), axis=0)*100))
            f.write("\nmax vertex error: " + str(np.mean(np.max(v_err, axis=1), axis=0)*100))
            f.write("\npampjpe: " + str(np.mean(pa_err)*100))

        # write errors to the excel
        sheet_loc_dict = {
            'lab1': ['A1', 'B1', 'A2', 'B2'],
            'lab2': ['C1', 'D1', 'C2', 'D2'],
            'furnished': ['E1', 'F1', 'E2', 'F2'],
            'rain': ['G1', 'H1', 'G2', 'H2'],
            'smoke': ['I1', 'J1', 'I2', 'J2'],
            'poor_lighting': ['K1', 'L1', 'K2', 'L2'],
            'occlusion': ['M1', 'N1', 'M2', 'N2'],
        }
        sheet_locs = sheet_loc_dict[args.test_scene]
        excel_path = os.path.join(args.output_dir, 'error', 'error.xlsx')
        if not os.path.exists(excel_path):
            work_book = Workbook()
            work_sheet = work_book.active
            work_sheet.title = "Sheet1"
        else:
            work_book = load_workbook(filename=excel_path)
            work_sheet = work_book['Sheet1']
        work_sheet[sheet_locs[0]] = str(np.mean(j_err)*100)
        work_sheet[sheet_locs[1]] = str(np.mean(v_err)*100)
        work_sheet[sheet_locs[2]] = str(np.mean(np.max(j_err, axis=1), axis=0)*100)
        work_sheet[sheet_locs[3]] = str(np.mean(np.max(v_err, axis=1), axis=0)*100)
        work_book.save(excel_path)

    return val_mPVE, val_mPJPE, val_PAmPJPE, val_count


def main(args):
    global logger
    # Setup CUDA, GPU & distributed training
    args.num_gpus = int(os.environ['WORLD_SIZE']) if 'WORLD_SIZE' in os.environ else 1
    os.environ['OMP_NUM_THREADS'] = str(args.num_workers)
    print('set os.environ[OMP_NUM_THREADS] to {}'.format(os.environ['OMP_NUM_THREADS']))

    args.distributed = args.num_gpus > 1
    torch.cuda.set_device(args.gpu_idx)
    args.device = torch.device(args.device)
    if args.distributed:
        # print("Init distributed training on local rank {} ({}), rank {}, world size {}".format(args.local_rank, int(os.environ["LOCAL_RANK"]), int(os.environ["NODE_RANK"]), args.num_gpus))
        torch.cuda.set_device(args.local_rank)
        torch.distributed.init_process_group(
            backend='nccl', init_method='env://'
        )
        local_rank = int(os.environ["LOCAL_RANK"])
        args.device = torch.device("cuda", local_rank)
        synchronize()

    mkdir(args.output_dir)
    logger = setup_logger("AMMFusion", args.output_dir, get_rank())
    set_seed(args.seed, args.num_gpus)
    logger.info("Using {} GPUs".format(args.num_gpus))
    args.inputs = args.inputs.replace(' ','').split(',')
    if args.model == 'TokenFusion':
        args.enabled_inputs = args.inputs
        args.inputs = ['image0','image1','depth0','depth1','radar0']
    args.input_dict = {}
    for m in ['image', 'depth', 'radar']:
        args.input_dict[m] = [i for i in args.inputs if m in i]
    if not args.inputs:
        raise RuntimeError("No input modality!")
    # Mesh and SMPL utils
    if args.mesh_type == 'smplx':
        smpl = SMPLX().to(args.device)
        mesh_sampler = SMPLXMesh()
        max_position_embeddings = 677
    elif args.mesh_type == 'smpl':
        smpl = SMPL().to(args.device)
        mesh_sampler = Mesh()
        max_position_embeddings = 455
    elif args.mesh_type == 'smplh36m':
        smpl = SMPLH36M().to(args.device)
        mesh_sampler = Mesh()
        max_position_embeddings = 445
    # Load model
    trans_encoder = []

    input_feat_dim = [int(item) for item in args.input_feat_dim.split(',')]
    hidden_feat_dim = [int(item) for item in args.hidden_feat_dim.split(',')]
    output_feat_dim = input_feat_dim[1:] + [args.output_dim]

    # which encoder block to have graph convs
    which_blk_graph = [int(item) for item in args.which_gcn.split(',')]
    
    start_epoch = 0
    if args.resume_checkpoint!=None and args.resume_checkpoint!='None':
        # if only run eval, load checkpoint
        logger.info("Evaluation: Loading from checkpoint {}".format(args.resume_checkpoint))
        _model = torch.load(os.path.join(args.resume_checkpoint, 'model.bin'), map_location='cpu')
        # for fine-tuning or resume training or inference, load weights from checkpoint
        logger.info("Loading state dict from checkpoint {}".format(args.resume_checkpoint))
        start_epoch = int(args.resume_checkpoint.split('-')[1])
        # workaround approach to load sparse tensor in graph conv.
        states = torch.load(os.path.join(args.resume_checkpoint, 'state_dict.bin'), map_location='cpu')
        # del checkpoint_loaded
        _model.load_state_dict(states, strict=False)
        del states
        gc.collect()
        torch.cuda.empty_cache()
    else:
        # init three transformer-encoder blocks in a loop
        for i in range(len(output_feat_dim)):
            config_class, model_class = BertConfig, Graphormer
            config = config_class.from_pretrained(args.config_name if args.config_name \
                    else args.model_name_or_path)

            config.device = args.device
            config.output_attentions = False
            config.hidden_dropout_prob = args.drop_out
            config.img_feature_dim = input_feat_dim[i] 
            config.output_feature_dim = output_feat_dim[i]
            args.hidden_size = hidden_feat_dim[i]
            args.intermediate_size = int(args.hidden_size*args.interm_size_scale)
            config.max_position_embeddings = max_position_embeddings

            if which_blk_graph[i]==1:
                config.graph_conv = True
                logger.info("Add Graph Conv")
            else:
                config.graph_conv = False

            config.mesh_type = args.mesh_type

            # update model structure if specified in arguments
            update_params = ['num_hidden_layers', 'hidden_size', 'num_attention_heads', 'intermediate_size']

            for param in update_params:
                arg_param = getattr(args, param)
                config_param = getattr(config, param)
                if arg_param > 0 and arg_param != config_param:
                    logger.info("Update config parameter {}: {} -> {}".format(param, config_param, arg_param))
                    setattr(config, param, arg_param)

            # init a transformer encoder and append it to a list
            assert config.hidden_size % config.num_attention_heads == 0
            model = model_class(config=config) 
            logger.info("Init model from scratch.")
            trans_encoder.append(model)

        
        # init ImageNet pre-trained backbone model
        if args.arch=='hrnet':
            hrnet_yaml = 'models/hrnet/cls_hrnet_w40_sgd_lr5e-2_wd1e-4_bs32_x100.yaml'
            hrnet_checkpoint = 'models/hrnet/hrnetv2_w40_imagenet_pretrained.pth'
            hrnet_update_config(hrnet_config, hrnet_yaml)
            img_backbone = get_cls_net_gridfeat(hrnet_config, pretrained=hrnet_checkpoint)
            logger.info('=> loading hrnet-v2-w40 model')
        elif args.arch=='hrnet-w64':
            hrnet_yaml = 'models/hrnet/cls_hrnet_w64_sgd_lr5e-2_wd1e-4_bs32_x100.yaml'
            hrnet_checkpoint = 'models/hrnet/hrnetv2_w64_imagenet_pretrained.pth'
            hrnet_update_config(hrnet_config, hrnet_yaml)
            img_backbone = get_cls_net_gridfeat(hrnet_config, pretrained=hrnet_checkpoint)
            logger.info('=> loading hrnet-v2-w64 model')
        else:
            print("=> using pre-trained model '{}'".format(args.arch))
            img_backbone = models.__dict__[args.arch](pretrained=True)
            # remove the last fc layer
            img_backbone = torch.nn.Sequential(*list(img_backbone.children())[:-2])

        trans_encoder = torch.nn.Sequential(*trans_encoder)
        total_params = sum(p.numel() for p in trans_encoder.parameters())
        logger.info('AMMFusion encoders total parameters: {}'.format(total_params))
        backbone_total_params = sum(p.numel() for p in img_backbone.parameters())
        logger.info('Backbone total parameters: {}'.format(backbone_total_params))
        
        if args.model == 'PointWImageFeat':
            mlps = [2048,4096,4096,2048]
        elif args.use_point_feat:
            mlps = [3,128,128,1024,1024,2048]
        else:
            mlps = [0,128,128,1024,1024,2048]
        radar_backbone = PointnetSAModule(npoint=args.num_clusters, radius=0.4, nsample=32, mlp=mlps.copy())
        depth_backbone = PointnetSAModule(npoint=49, radius=0.4, nsample=64, mlp=mlps.copy())
        
        backbone = dict(radar=radar_backbone, image=img_backbone, depth=depth_backbone)

        # build end-to-end AMMfusion network (backbone + multi-layer Fusion Transformer)
        Model = getattr(Models, args.model)
        _model = Model(args, backbone, trans_encoder)

    # update configs to enable attention outputs
    if args.show_att:
        setattr(_model.trans_encoder[-1].config,'output_attentions', True)
        setattr(_model.trans_encoder[-1].config,'output_hidden_states', True)
        _model.trans_encoder[-1].bert.encoder.output_attentions = True
        _model.trans_encoder[-1].bert.encoder.output_hidden_states =  True
        for iter_layer in range(4):
            _model.trans_encoder[-1].bert.encoder.layer[iter_layer].attention.self.output_attentions = True
        setattr(_model.trans_encoder[-1].config,'device', args.device)

    _model.to(args.device)
    logger.info("Training parameters %s", args)
    Dataset = getattr(Datasets, args.dataset)
    dataset = Dataset(args)

    if args.train==True:
        if args.mix_data:
            datasets = [
                dataset,
                Datasets.CLIFFDataset(args, data_path='/remote-home/chenanjun/dataset/COCO'),
                Datasets.CLIFFDataset(args, data_path='/remote-home/chenanjun/dataset/mpii'),
            ]
            dataset = torch.utils.data.ConcatDataset(datasets)
        if args.eval_test_dataset:
            dataset_train = dataset
            eval_args = deepcopy(args)
            eval_args.train = False
            dataset_eval = Dataset(eval_args)
        else:
            train_size = int(0.9 * len(dataset))
            eval_size = len(dataset) - train_size
            dataset_train, dataset_eval = torch.utils.data.random_split(dataset, [train_size, eval_size])
        train_dataloader = make_data_loader(args, dataset_train, args.distributed, is_train=True, 
                                            scale_factor=args.img_scale_factor)
        val_dataloader = make_data_loader(args, dataset_eval, args.distributed, is_train=False, 
                                            scale_factor=args.img_scale_factor)
        run_train(args, train_dataloader, val_dataloader, _model, smpl, mesh_sampler, start_epoch)
        
    else:
        val_dataloader = make_data_loader(args, dataset, args.distributed, is_train=False, scale_factor=args.img_scale_factor)
        run_eval_general(args, val_dataloader, _model, smpl, mesh_sampler)


def parse_args():
    parser = argparse.ArgumentParser()
    #########################################################
    # Data related arguments
    #########################################################
    parser.add_argument("--data_path", default='/home/nesc525/drivers/6/mmBody', type=str, required=False,
                        help="Directory with all datasets, each in one subfolder")
    parser.add_argument("--num_workers", default=4, type=int, 
                        help="Workers in dataloader.")
    parser.add_argument("--img_scale_factor", default=1, type=int, 
                        help="adjust image resolution.") 
    parser.add_argument("--test_scene", type=str, default='lab1')
    parser.add_argument("--seq_idxes", type=str, default='') 
    parser.add_argument('--skip_head', type=int, default=0)
    parser.add_argument('--skip_tail', type=int, default=0)
    parser.add_argument('--dataset', type=str, default='mmBodyDataset')
    parser.add_argument('--num_points', type=int, default=1024)
    parser.add_argument('--add_rgb', action="store_true", help='add rgb values')
    parser.add_argument('--inputs', type=str, default='image0,radar', help='input data')
    parser.add_argument('--trans_coor_to_cam', action="store_true")
    parser.add_argument("--mesh_type", default='smplx', type=str, help="smplx or smpl") 
    parser.add_argument('--mask_ratio', type=float, default=0.3)
    parser.add_argument('--num_clusters', type=int, default=49)
    parser.add_argument('--mask_limbs', action="store_true")
    parser.add_argument('--point_mask_ratio', type=float, default=0.99)
    parser.add_argument('--num_mask_limbs', type=int, default=None)
    parser.add_argument('--need_augm', action="store_true")
    parser.add_argument('--mix_data', action="store_true")
    parser.add_argument('--eval_test_dataset', action="store_true")
    #########################################################
    # Loading/saving checkpoints
    #########################################################
    parser.add_argument("--model_name_or_path", default='src/modeling/bert/bert-base-uncased/', type=str, required=False,
                        help="Path to pre-trained transformer model or model type.")
    parser.add_argument("--resume_checkpoint", default=None, type=str, required=False,
                        help="Path to specific checkpoint for resume training.")
    parser.add_argument("--output_dir", default='output/output', type=str, required=False,
                        help="The output directory to save checkpoint and test results.")
    parser.add_argument("--config_name", default="", type=str, 
                        help="Pretrained config name or path if not the same as model_name.")
    #########################################################
    # Training parameters
    #########################################################
    parser.add_argument("--per_gpu_train_batch_size", default=10, type=int, 
                        help="Batch size per GPU/CPU for training.")
    parser.add_argument("--per_gpu_eval_batch_size", default=10, type=int, 
                        help="Batch size per GPU/CPU for evaluation.")
    parser.add_argument('--lr', "--learning_rate", default=1e-4, type=float, 
                        help="The initial lr.")
    parser.add_argument("--num_train_epochs", default=50, type=int, 
                        help="Total number of training epochs to perform.")
    parser.add_argument("--vertices_loss_weight", default=100.0, type=float)
    parser.add_argument("--joints_loss_weight", default=1000.0, type=float)
    parser.add_argument("--score_loss_weight", default=100.0, type=float)
    parser.add_argument("--vloss_w_full", default=0.33, type=float) 
    parser.add_argument("--vloss_w_sub", default=0.33, type=float) 
    parser.add_argument("--vloss_w_sub2", default=0.33, type=float) 
    parser.add_argument("--drop_out", default=0.1, type=float, 
                        help="Drop out ratio in BERT.")
    #########################################################
    # Model architectures
    #########################################################
    parser.add_argument('-a', '--arch', default='hrnet-w64',
                    help='CNN backbone architecture: hrnet-w64, hrnet, resnet50')
    parser.add_argument("--model", default='AdaptiveFusion', type=str,
                        help='Choose the model')
    parser.add_argument("--num_hidden_layers", default=4, type=int, required=False, 
                        help="Update model config if given")
    parser.add_argument("--hidden_size", default=-1, type=int, required=False, 
                        help="Update model config if given")
    parser.add_argument("--num_attention_heads", default=4, type=int, required=False, 
                        help="Update model config if given. Note that the division of "
                        "hidden_size / num_attention_heads should be in integer.")
    parser.add_argument("--intermediate_size", default=-1, type=int, required=False, 
                        help="Update model config if given.")
    parser.add_argument("--input_feat_dim", default='2051,512,128', type=str, 
                        help="The Image Feature Dimension.")          
    parser.add_argument("--hidden_feat_dim", default='1024,256,64', type=str, 
                        help="The Image Feature Dimension.")   
    parser.add_argument("--which_gcn", default='0,0,1', type=str, 
                        help="which encoder block to have graph conv. Encoder1, Encoder2, Encoder3. Default: only Encoder3 has graph conv") 
    parser.add_argument("--interm_size_scale", default=2, type=int)
    parser.add_argument('--use_point_feat', action="store_true", help='use point feature')
    #########################################################
    # Others
    #########################################################
    parser.add_argument('--train', dest="train", action="store_true", help='train or test')
    parser.add_argument('--visual', dest="visual", action="store_true", help='visual')
    parser.add_argument('--pause_at_start', action="store_true")
    parser.add_argument('--logging_steps', type=int, default=1000, 
                        help="Log every X steps.")
    parser.add_argument("--device", type=str, default='cuda', 
                        help="cuda or cpu")
    parser.add_argument("--gpu_idx", type=int, default=0, help="select gpu")
    parser.add_argument('--seed', type=int, default=88, 
                        help="random seed for initialization.")
    parser.add_argument("--local_rank", type=int, default=0, 
                        help="For distributed training.")
    parser.add_argument('--max_num_batch', type=int, default=10000)
    parser.add_argument('--save_snapshot', action="store_true")
    parser.add_argument('--output_dim', type=int, default=3)
    parser.add_argument('--points_w_image_feat', action="store_true")
    parser.add_argument('--fix_modalities', action="store_true")
    parser.add_argument('--wo_GIM', action="store_true")
    parser.add_argument('--wo_MMM', action="store_true")
    parser.add_argument('--wo_local_feat', action="store_true")
    parser.add_argument('--show_att', action="store_true")
    parser.add_argument('--joint_id', type=int, default=0)
    parser.add_argument('--calib_emb', action="store_true")
    parser.add_argument('--shuffle_inputs', action="store_true")
    parser.add_argument('--joints_2d_loss', type=str, default='')

    args = parser.parse_args()
    return args

if __name__ == "__main__":
    args = parse_args()
    main(args)
